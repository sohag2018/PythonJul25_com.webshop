import openpyxl

class ExcelReader:
#method to return rowCount
    def getRowCount(file_path,sheetname):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheetname]
        return (sheet.max_row)

#method to return colCount
    def getColCount(file_path,sheetname):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheetname]
        return (sheet.max_column)

#method to return cellData
    def getCellData(file_path,sheetname,rowNum,colNum):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheetname]
        return sheet.cell(row=rowNum, column=colNum).value



